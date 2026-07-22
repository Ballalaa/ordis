"""
Ordis backend — serves the static site, stores signup submissions in a
SQLite database (mirrored into an Excel export), and powers an admin
dashboard for managing members, pricing plans, and basic analytics.

Run with:  python3 server.py
Site:      http://localhost:8000/home
Admin:     http://localhost:8000/admin-dashboard  (default login printed on first run)
"""
import json
import os
import re
import secrets
import sqlite3
from datetime import datetime, timedelta, timezone
from functools import wraps

from flask import Flask, jsonify, redirect, request, send_from_directory, session
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "ordis.db")
XLSX_PATH = os.path.join(DATA_DIR, "registrations.xlsx")
SECRET_KEY_PATH = os.path.join(DATA_DIR, "secret_key.txt")

EXCEL_HEADERS = ["ID", "სახელი", "გვარი", "ელ. ფოსტა", "ტელეფონი", "პაკეტი", "სტატუსი", "დარეგისტრირდა"]

NAME_RE = re.compile(r"^[A-Za-zა-ჰ' -]{2,60}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
# Georgian mobile: 9 digits starting with 5, optionally prefixed with +995/995
PHONE_RE = re.compile(r"^5\d{8}$")

STATUSES = ["new", "contacted", "active", "inactive"]
STATUS_LABELS = {"new": "ახალი", "contacted": "დაკონტაქტებული", "active": "აქტიური", "inactive": "არააქტიური"}

DEFAULT_ADMIN_USER = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"

DEFAULT_PLANS = [
    dict(
        name="სტანდარტული", price="0", currency="", period="", format_label="უფასოდ სამუდამოდ",
        description="ძირითადი სერვისები უფასოდ. მცირე ჯგუფებისთვის, ვისაც სურს აღრიცხვის მოწესრიგება დამატებითი ხარჯის გარეშე.",
        is_popular=0, sort_order=1,
        features=["მოსწავლეთა და ჯგუფების აღრიცხვა", "ჯგუფების განრიგი", "დავალიანების ავტომატური დარიცხვა",
                  "ჩარიცხვების აღრიცხვა", "საბაზისო ანალიტიკა"],
    ),
    dict(
        name="Plus", price="1", currency="₾", period="თვე", format_label="თითო მოსწავლეზე",
        description="სტანდარტულზე მეტი სერვისი, დამატებით საკომუნიკაციო მოდული. საშუალო და დიდი ზომის აკადემიებისთვის.",
        is_popular=1, sort_order=2,
        features=["სტანდარტულის ყველა შესაძლებლობა", "SMS და აპლიკაციის შეტყობინებები",
                  "შეღავათებისა და ფასდაკლების მართვა", "გაფართოებული ანალიტიკა", "პრიორიტეტული მხარდაჭერა"],
    ),
    dict(
        name="Business", price="მორგებული", currency="", period="", format_label="ინდივიდუალური შეთავაზება",
        description="მრავალფილიალიანი ან კომპლექსური სტრუქტურის ორგანიზაციებისთვის.",
        is_popular=0, sort_order=3,
        features=["Plus-ის ყველა შესაძლებლობა", "რამდენიმე ფილიალის მართვა", "პერსონალური ონბორდინგი",
                  "API და ინტეგრაციები", "გამოყოფილი მენეჯერი"],
    ),
]

app = Flask(__name__, static_folder=None)


def get_secret_key():
    if os.path.exists(SECRET_KEY_PATH):
        with open(SECRET_KEY_PATH, "r", encoding="utf-8") as f:
            return f.read().strip()
    key = secrets.token_hex(32)
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(SECRET_KEY_PATH, "w", encoding="utf-8") as f:
        f.write(key)
    return key


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT NOT NULL,
            plan TEXT,
            status TEXT NOT NULL DEFAULT 'new',
            created_at TEXT NOT NULL
        )
        """
    )
    # Migrate databases created before plan/status existed.
    existing_cols = {row["name"] for row in conn.execute("PRAGMA table_info(registrations)")}
    if "plan" not in existing_cols:
        conn.execute("ALTER TABLE registrations ADD COLUMN plan TEXT")
    if "status" not in existing_cols:
        conn.execute("ALTER TABLE registrations ADD COLUMN status TEXT NOT NULL DEFAULT 'new'")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            price TEXT NOT NULL,
            currency TEXT DEFAULT '',
            period TEXT DEFAULT '',
            format_label TEXT DEFAULT '',
            description TEXT DEFAULT '',
            features TEXT NOT NULL DEFAULT '[]',
            is_popular INTEGER NOT NULL DEFAULT 0,
            sort_order INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        )
        """
    )

    if conn.execute("SELECT COUNT(*) FROM plans").fetchone()[0] == 0:
        for p in DEFAULT_PLANS:
            conn.execute(
                "INSERT INTO plans (name, price, currency, period, format_label, description, "
                "features, is_popular, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (p["name"], p["price"], p["currency"], p["period"], p["format_label"], p["description"],
                 json.dumps(p["features"], ensure_ascii=False), p["is_popular"], p["sort_order"]),
            )

    if conn.execute("SELECT COUNT(*) FROM admin_users").fetchone()[0] == 0:
        conn.execute(
            "INSERT INTO admin_users (username, password_hash) VALUES (?, ?)",
            (DEFAULT_ADMIN_USER, generate_password_hash(DEFAULT_ADMIN_PASSWORD, method="pbkdf2:sha256")),
        )
        print(
            f"\n[Ordis] Created default admin login — username: {DEFAULT_ADMIN_USER}  "
            f"password: {DEFAULT_ADMIN_PASSWORD}\n[Ordis] Change it from the admin dashboard's Settings tab.\n"
        )

    conn.commit()
    return conn


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_id"):
            return jsonify({"ok": False, "error": "unauthorized"}), 401
        return view(*args, **kwargs)

    return wrapped


def plan_row_to_dict(row):
    return {
        "id": row["id"],
        "name": row["name"],
        "price": row["price"],
        "currency": row["currency"],
        "period": row["period"],
        "format_label": row["format_label"],
        "description": row["description"],
        "features": json.loads(row["features"]),
        "is_popular": bool(row["is_popular"]),
        "sort_order": row["sort_order"],
    }


def member_row_to_dict(row):
    return {
        "id": row["id"],
        "first_name": row["first_name"],
        "last_name": row["last_name"],
        "email": row["email"],
        "phone": row["phone"],
        "plan": row["plan"],
        "status": row["status"],
        "status_label": STATUS_LABELS.get(row["status"], row["status"]),
        "created_at": row["created_at"],
    }


def append_to_excel(row_id, first_name, last_name, email, phone, plan, status, created_at):
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        ws.append(EXCEL_HEADERS)
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    ws.append([row_id, first_name, last_name, email, "+995" + phone, plan or "", STATUS_LABELS.get(status, status), created_at])
    wb.save(XLSX_PATH)


def normalize_phone(raw):
    digits = re.sub(r"\D", "", raw or "")
    if digits.startswith("995") and len(digits) == 12:
        digits = digits[3:]
    return digits


# ---------------------------------------------------------------- public API

@app.post("/api/register")
def register():
    payload = request.get_json(silent=True) or {}

    first_name = (payload.get("first_name") or "").strip()
    last_name = (payload.get("last_name") or "").strip()
    email = (payload.get("email") or "").strip().lower()
    phone = normalize_phone(payload.get("phone"))
    consent = bool(payload.get("consent"))
    plan = (payload.get("plan") or "").strip() or None

    errors = {}
    if not NAME_RE.match(first_name):
        errors["first_name"] = "სახელი უნდა იყოს 2-60 სიმბოლო."
    if not NAME_RE.match(last_name):
        errors["last_name"] = "გვარი უნდა იყოს 2-60 სიმბოლო."
    if not EMAIL_RE.match(email):
        errors["email"] = "მიუთითეთ ვალიდური ელ. ფოსტა."
    if not PHONE_RE.match(phone):
        errors["phone"] = "მიუთითეთ ვალიდური ქართული მობილურის ნომერი (5XXXXXXXX)."
    if not consent:
        errors["consent"] = "აუცილებელია წესებსა და პირობებზე დათანხმება."

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    created_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO registrations (first_name, last_name, email, phone, plan, status, created_at) "
            "VALUES (?, ?, ?, ?, ?, 'new', ?)",
            (first_name, last_name, email, phone, plan, created_at),
        )
        conn.commit()
        row_id = cur.lastrowid
    finally:
        conn.close()

    try:
        append_to_excel(row_id, first_name, last_name, email, phone, plan, "new", created_at)
    except PermissionError:
        app.logger.warning(
            "Could not write to %s (file may be open in Excel). Row %s is saved in the database only.",
            XLSX_PATH, row_id,
        )

    return jsonify({"ok": True, "id": row_id})


@app.get("/api/plans")
def public_plans():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM plans ORDER BY sort_order ASC").fetchall()
    finally:
        conn.close()
    return jsonify({"ok": True, "plans": [plan_row_to_dict(r) for r in rows]})


# ---------------------------------------------------------------- admin auth

@app.post("/admin/api/login")
def admin_login():
    payload = request.get_json(silent=True) or {}
    username = (payload.get("username") or "").strip()
    password = payload.get("password") or ""

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM admin_users WHERE username = ?", (username,)).fetchone()
    finally:
        conn.close()

    if not row or not check_password_hash(row["password_hash"], password):
        return jsonify({"ok": False, "error": "მომხმარებელი ან პაროლი არასწორია."}), 401

    session["admin_id"] = row["id"]
    session["admin_username"] = row["username"]
    return jsonify({"ok": True})


@app.post("/admin/api/logout")
def admin_logout():
    session.clear()
    return jsonify({"ok": True})


@app.get("/admin/api/me")
def admin_me():
    if not session.get("admin_id"):
        return jsonify({"ok": False}), 401
    return jsonify({"ok": True, "username": session.get("admin_username")})


@app.post("/admin/api/change-password")
@login_required
def admin_change_password():
    payload = request.get_json(silent=True) or {}
    current = payload.get("current_password") or ""
    new = payload.get("new_password") or ""

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM admin_users WHERE id = ?", (session["admin_id"],)).fetchone()
        if not row or not check_password_hash(row["password_hash"], current):
            return jsonify({"ok": False, "error": "მიმდინარე პაროლი არასწორია."}), 400
        if len(new) < 8:
            return jsonify({"ok": False, "error": "ახალი პაროლი უნდა იყოს მინიმუმ 8 სიმბოლო."}), 400
        conn.execute(
            "UPDATE admin_users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new, method="pbkdf2:sha256"), row["id"]),
        )
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True})


# ------------------------------------------------------------ admin: members

@app.get("/admin/api/members")
@login_required
def admin_list_members():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    plan = (request.args.get("plan") or "").strip()

    clauses, params = [], []
    if q:
        clauses.append("(first_name LIKE ? OR last_name LIKE ? OR email LIKE ? OR phone LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like, like]
    if status:
        clauses.append("status = ?")
        params.append(status)
    if plan:
        clauses.append("plan = ?")
        params.append(plan)

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT * FROM registrations {where} ORDER BY created_at DESC", params
        ).fetchall()
    finally:
        conn.close()

    return jsonify({"ok": True, "members": [member_row_to_dict(r) for r in rows]})


@app.patch("/admin/api/members/<int:member_id>")
@login_required
def admin_update_member(member_id):
    payload = request.get_json(silent=True) or {}
    fields, params = [], []

    if "status" in payload:
        if payload["status"] not in STATUSES:
            return jsonify({"ok": False, "error": "არასწორი სტატუსი."}), 400
        fields.append("status = ?")
        params.append(payload["status"])

    if "plan" in payload:
        fields.append("plan = ?")
        params.append(payload["plan"] or None)

    if not fields:
        return jsonify({"ok": False, "error": "არაფერია განსახლებადი."}), 400

    params.append(member_id)

    conn = get_db()
    try:
        conn.execute(f"UPDATE registrations SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
        row = conn.execute("SELECT * FROM registrations WHERE id = ?", (member_id,)).fetchone()
    finally:
        conn.close()

    if not row:
        return jsonify({"ok": False, "error": "მომხმარებელი ვერ მოიძებნა."}), 404

    return jsonify({"ok": True, "member": member_row_to_dict(row)})


@app.delete("/admin/api/members/<int:member_id>")
@login_required
def admin_delete_member(member_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM registrations WHERE id = ?", (member_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True})


# -------------------------------------------------------------- admin: plans

@app.get("/admin/api/plans")
@login_required
def admin_list_plans():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM plans ORDER BY sort_order ASC").fetchall()
    finally:
        conn.close()
    return jsonify({"ok": True, "plans": [plan_row_to_dict(r) for r in rows]})


def _plan_fields_from_payload(payload):
    return dict(
        name=(payload.get("name") or "").strip(),
        price=(payload.get("price") or "").strip(),
        currency=(payload.get("currency") or "").strip(),
        period=(payload.get("period") or "").strip(),
        format_label=(payload.get("format_label") or "").strip(),
        description=(payload.get("description") or "").strip(),
        features=[f.strip() for f in (payload.get("features") or []) if f.strip()],
        is_popular=1 if payload.get("is_popular") else 0,
        sort_order=int(payload.get("sort_order") or 0),
    )


@app.post("/admin/api/plans")
@login_required
def admin_create_plan():
    payload = request.get_json(silent=True) or {}
    data = _plan_fields_from_payload(payload)

    if not data["name"] or not data["price"]:
        return jsonify({"ok": False, "error": "სახელი და ფასი სავალდებულოა."}), 400

    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO plans (name, price, currency, period, format_label, description, "
            "features, is_popular, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (data["name"], data["price"], data["currency"], data["period"], data["format_label"],
             data["description"], json.dumps(data["features"], ensure_ascii=False),
             data["is_popular"], data["sort_order"]),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM plans WHERE id = ?", (cur.lastrowid,)).fetchone()
    finally:
        conn.close()

    return jsonify({"ok": True, "plan": plan_row_to_dict(row)})


@app.put("/admin/api/plans/<int:plan_id>")
@login_required
def admin_update_plan(plan_id):
    payload = request.get_json(silent=True) or {}
    data = _plan_fields_from_payload(payload)

    if not data["name"] or not data["price"]:
        return jsonify({"ok": False, "error": "სახელი და ფასი სავალდებულოა."}), 400

    conn = get_db()
    try:
        conn.execute(
            "UPDATE plans SET name=?, price=?, currency=?, period=?, format_label=?, description=?, "
            "features=?, is_popular=?, sort_order=? WHERE id=?",
            (data["name"], data["price"], data["currency"], data["period"], data["format_label"],
             data["description"], json.dumps(data["features"], ensure_ascii=False),
             data["is_popular"], data["sort_order"], plan_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM plans WHERE id = ?", (plan_id,)).fetchone()
    finally:
        conn.close()

    if not row:
        return jsonify({"ok": False, "error": "პაკეტი ვერ მოიძებნა."}), 404

    return jsonify({"ok": True, "plan": plan_row_to_dict(row)})


@app.delete("/admin/api/plans/<int:plan_id>")
@login_required
def admin_delete_plan(plan_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM plans WHERE id = ?", (plan_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True})


# --------------------------------------------------------- admin: analytics

@app.get("/admin/api/analytics")
@login_required
def admin_analytics():
    conn = get_db()
    try:
        total = conn.execute("SELECT COUNT(*) FROM registrations").fetchone()[0]

        now = datetime.now(timezone.utc)
        cutoff_7 = (now - timedelta(days=7)).isoformat(timespec="seconds")
        cutoff_30 = (now - timedelta(days=30)).isoformat(timespec="seconds")
        new_7d = conn.execute(
            "SELECT COUNT(*) FROM registrations WHERE created_at >= ?", (cutoff_7,)
        ).fetchone()[0]
        new_30d = conn.execute(
            "SELECT COUNT(*) FROM registrations WHERE created_at >= ?", (cutoff_30,)
        ).fetchone()[0]

        by_status_rows = conn.execute(
            "SELECT status, COUNT(*) as n FROM registrations GROUP BY status"
        ).fetchall()
        by_status = {s: 0 for s in STATUSES}
        for r in by_status_rows:
            by_status[r["status"]] = r["n"]

        by_plan_rows = conn.execute(
            "SELECT COALESCE(plan, '') as plan, COUNT(*) as n FROM registrations GROUP BY plan"
        ).fetchall()
        by_plan = {(r["plan"] or "დაუზუსტებელი"): r["n"] for r in by_plan_rows}

        daily_rows = conn.execute(
            "SELECT substr(created_at, 1, 10) as day, COUNT(*) as n FROM registrations "
            "WHERE created_at >= ? GROUP BY day ORDER BY day ASC",
            ((now - timedelta(days=13)).isoformat(timespec="seconds"),),
        ).fetchall()
        daily_map = {r["day"]: r["n"] for r in daily_rows}
        daily = []
        for i in range(13, -1, -1):
            day = (now - timedelta(days=i)).date().isoformat()
            daily.append({"date": day, "count": daily_map.get(day, 0)})

        plans_count = conn.execute("SELECT COUNT(*) FROM plans").fetchone()[0]
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "total_members": total,
        "new_7d": new_7d,
        "new_30d": new_30d,
        "by_status": by_status,
        "status_labels": STATUS_LABELS,
        "by_plan": by_plan,
        "daily": daily,
        "plans_count": plans_count,
    })


# ---------------------------------------------------------------- page routes

@app.get("/admin-dashboard")
def admin_page():
    if not session.get("admin_id"):
        return send_from_directory(BASE_DIR, "admin-login.html")
    return send_from_directory(BASE_DIR, "admin.html")


@app.get("/admin")
def admin_redirect():
    return redirect("/admin-dashboard")


@app.get("/home")
def home():
    return send_from_directory(BASE_DIR, "index.html")


@app.get("/")
def index():
    return redirect("/home")


@app.get("/<path:filename>")
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)


# These run on import too (not just `python3 server.py` directly) — this
# matters when a WSGI host like PythonAnywhere imports `app` from this file
# without ever executing the __main__ block below.
os.makedirs(DATA_DIR, exist_ok=True)
app.secret_key = get_secret_key()
get_db().close()  # ensure tables + seed data exist before serving

if __name__ == "__main__":
    # use_reloader=False: the debug auto-reloader spawns a child process that
    # doesn't match simple `pkill -f server.py` patterns and can linger,
    # leaving stale processes fighting over the port after restarts.
    app.run(host="127.0.0.1", port=8000, debug=True, use_reloader=False)
